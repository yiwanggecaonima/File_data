import pymongo

client = pymongo.MongoClient('127.0.0.1',27017)
db = client['天眼查']
import openpyxl

# 存入xlsx
outwb = openpyxl.Workbook()  # 打开一个将写的文件
outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
# 维护表头
tableTitle = ['公司名称','联系人','电话号码','地区','注册资金','成立时间','电子邮箱','状态','链接地址']
for col in range(len(tableTitle)):
        c = col + 1
        outws.cell(row=1, column=c).value = tableTitle[col]
        
datas = db['咨询'].find({},{'_id':0})
try:
    for row in range(1,712563):
        print('---------',row,'-----------')
        for key, value in datas[row].items():
            if key == '企业名称':
                print(value)
                outws.cell(row, 1).value = value  # 写文件
            elif key == '法人代表':
                outws.cell(row, 2).value = value
            elif key == '联系电话':
                outws.cell(row, 3).value = value
            elif key == '区域':
                outws.cell(row, 4).value = value
            elif key == '注册资金':
                outws.cell(row, 5).value = value
            elif key == '成立时间':
                outws.cell(row, 6).value = value
            elif key == '邮箱':
                outws.cell(row,7).value = value
            elif key == '状态':
                outws.cell(row,8).value = value
            elif key == '网址链接':
                outws.cell(row,9).value = value
except Exception:
    pass


saveExcel = "咨询行业.xls"
outwb.save(saveExcel)  # 保存





# 读取xlsx
#!/usr/bin/env python 
# -*- coding: utf-8 -*-

from openpyxl.reader.excel import load_workbook
import json

# 读取excel2007文件
wb = load_workbook(filename='/home/parrot/Desktop/666.xlsx')

# 显示有多少张表
print("Worksheet range(s):", wb.get_named_ranges())
print("Worksheet name(s):", wb.get_sheet_names())

# 取第一张表
sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetnames[0])

# 显示表名，表行数，表列数
print("Work Sheet Titile:", ws.title)
print("Work Sheet Rows:", ws.max_row)
print("Work Sheet Cols:", ws.max_column)



# 把数据存到字典中
for rx in range(2, 360):
    d ={} 
    pid = rx
    dengji = ws.cell(row=rx, column=1).value
    
    baidu = ws.cell(row=rx, column=2).value
    
    mafengwo_name = ws.cell(row=rx, column=5).value
    
    

    
    d["111"] = dengji
    d["222"] = 
    d["333"] = 
    d["444"] = 

    print(d)
    with open("/home/parrot/Desktop/666.txt","a") as f:
        f.write(json.dumps(d, ensure_ascii=False) + "\n")
# print('Total:%d' % len(data_dic))
# print(json.dumps(d, ensure_ascii=False))
